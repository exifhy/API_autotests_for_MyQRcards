from urllib import parse
import allure
import requests
from loguru import logger
from requests import JSONDecodeError
from utils.helper import Helper
from services.export.users.payloads import Payloads
from services.export.users.endpoints import Endpoints
from config.headers import Headers
from services.export.users.models.users_model import *
import time
from http import HTTPStatus
from dotenv import load_dotenv
import os
from faker import Faker
from openpyxl import load_workbook
from io import BytesIO


fake_ru = Faker('ru_RU')

load_dotenv()
API_TOKEN = os.getenv('API_TOKEN')


class ExportUsersAPI(Helper):

    def __init__(self):
        super().__init__()
        self.payloads = Payloads()
        self.endpoints = Endpoints()
        self.headers = Headers()

    @allure.step("Exports the list of users taking into account the specified filters (Customers).")
    def get_export_list_customers_by_user_id(self, user_id: int, name: str, surname: str, email: str, phone: str):
        params = {
            "userID": user_id,
            "isCustomer": "true",
            "includeTaskActuality": "true",
            "isDeleted": "false"
        }
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_list_users_endpoint, params=params,
            headers=self.headers.export_header(API_TOKEN)
        )
        end = time.time()
        logger.info(response.headers)
        logger.warning(response.request.url)
        self.attach_time(start, end)
        try:
            logger.warning(response.json())
        except JSONDecodeError:
            logger.warning("Received response is not a valid JSON")
        assert response.status_code == HTTPStatus.OK, f'Status code {response.status_code}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        # СОХРАНЕНИЕ ЛОКАЛЬНО EXCEL
        # output_file_path = "exported_data.xlsx"
        # workbook.save(output_file_path)

        sheet = workbook.active
        sheet_name = workbook.sheetnames

        # ВЫВОД СТРОК EXCEL ФАЙЛА
        # count = 0
        # for row in sheet.iter_rows(values_only=True):
        #     logger.warning(row)
        #     count += 1
        #     if count >= 10:
        #         break

        assert 'Пользователи' in sheet_name
        assert sheet['B3'].value == 'Фамилия*'
        assert sheet['B4'].value == surname
        assert sheet['C3'].value == 'Имя*'
        assert sheet['C4'].value == name
        assert sheet['D3'].value == 'Отчество'
        assert sheet['F3'].value == 'Пол'
        assert sheet['G3'].value == 'Телефон*'
        assert sheet['G4'].value == phone
        assert sheet['H3'].value == 'Электронная почта*'
        assert sheet['H3'].value == email

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Пользователи.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully export of list customers by userID..')

    @allure.step("Exports the list of users taking into account the specified filters (Employee).")
    def get_export_list_employee_by_user_id(self, user_id: int, name: str, surname: str, email: str, phone: str):
        params = {
            "userID": user_id,
            "isCustomer": "false",
            "isTeam": "false",
            "isBanned": "false",
            "isDeleted": "false",
            "isInitial": "false"
        }
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_list_users_endpoint, params=params,
            headers=self.headers.export_header(API_TOKEN)
        )
        end = time.time()
        logger.info(response.headers)
        logger.warning(response.request.url)
        self.attach_time(start, end)
        try:
            logger.warning(response.json())
        except JSONDecodeError:
            logger.warning("Received response is not a valid JSON")
        assert response.status_code == HTTPStatus.OK, f'Status code {response.status_code}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        # СОХРАНЕНИЕ ЛОКАЛЬНО EXCEL
        # output_file_path = "exported_data.xlsx"
        # workbook.save(output_file_path)

        sheet = workbook.active
        sheet_name = workbook.sheetnames

        # ВЫВОД СТРОК EXCEL ФАЙЛА
        # count = 0
        # for row in sheet.iter_rows(values_only=True):
        #     logger.warning(row)
        #     count += 1
        #     if count >= 10:
        #         break

        assert 'Пользователи' in sheet_name
        assert sheet['B3'].value == 'Фамилия*'
        assert sheet['B4'].value == surname
        assert sheet['C3'].value == 'Имя*'
        assert sheet['C4'].value == name
        assert sheet['D3'].value == 'Отчество'
        assert sheet['F3'].value == 'Пол'
        assert sheet['G3'].value == 'Телефон*'
        assert sheet['G4'].value == phone
        assert sheet['H3'].value == 'Электронная почта*'
        assert sheet['H3'].value == email

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Пользователи.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully export of list employee by userID.')
