from urllib import parse
import allure
import requests
from loguru import logger
from requests import JSONDecodeError
from utils.helper import Helper
from services.export.users.payloads import Payloads
from services.export.users.endpoints import Endpoints
from config.headers import Headers
from services.export.users.models.export_users_model import *
import time
from http import HTTPStatus
from dotenv import load_dotenv
import os
from faker import Faker
from openpyxl import load_workbook
from io import BytesIO


fake_ru = Faker('ru_RU')

load_dotenv()
API_TOKEN = os.getenv('API_TOKEN')


class ExportUsersAPI(Helper):

    def __init__(self):
        super().__init__()
        self.payloads = Payloads()
        self.endpoints = Endpoints()
        self.headers = Headers()

    @allure.step("Exports the list of users taking into account the specified filters (Customers).")
    def get_export_list_customers_by_user_id(
            self,
            user_id: int,
            name: str,
            surname: str,
            email: str,
            phone: str,
            role: str,
            district_name: str,
            company_name: str
    ):
        params = {
            "userID": user_id,
            "isCustomer": "true",
            "includeTaskActuality": "true",
            "isDeleted": "false"
        }
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_list_users_endpoint, params=params,
            headers=self.headers.export_header(API_TOKEN)
        )
        end = time.time()
        logger.info(response.headers)
        self.attach_time(start, end)
        self.attach_url(response.request.url)
        try:
            self.attach_response(response.json())
        except JSONDecodeError:
            logger.warning("Received response is not a valid JSON")
        assert response.status_code == HTTPStatus.OK, f'{response.status_code}, {response.json()}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        # СОХРАНЕНИЕ ЛОКАЛЬНО EXCEL
        # output_file_path = "exported_data.xlsx"
        # workbook.save(output_file_path)

        sheet = workbook.active
        sheet_name = workbook.sheetnames

        # ВЫВОД СТРОК EXCEL ФАЙЛА
        # count = 0
        # for row in sheet.iter_rows(values_only=True):
        #     logger.warning(row)
        #     count += 1
        #     if count >= 10:
        #         break

        assert 'Пользователи' in sheet_name
        assert sheet['A3'].value == 'Фамилия*', f'Expected Фамилия*, but got {sheet['A3'].value}'
        assert sheet['A4'].value == surname.strip(), f'Expected {surname.strip()}, but got {sheet['A4'].value}'
        assert sheet['B3'].value == 'Имя*', f'Expected Имя*, but got {sheet['B3'].value}'
        assert sheet['B4'].value == name.strip(), f'Expected {name.strip()}, but got {sheet['B4'].value}'
        assert sheet['C3'].value == 'Отчество', f'Expected Отчество, but got {sheet['C3'].value}'
        assert sheet['D3'].value == 'Пол', f'Expected Пол, but got {sheet['D3'].value}'
        assert sheet['E3'].value == 'Телефон*', f'Expected Телефон*, but got {sheet['E3'].value}'
        assert f"+{sheet['E4'].value}" == phone.strip(), f'Expected {phone.strip()}, but got +{sheet['E4'].value}'
        assert sheet['F3'].value == 'Электронная почта*', f'Expected Электронная почта*, but got {sheet['F3'].value}'
        assert sheet['F4'].value == email.strip(), f'Expected {email.strip()}, but got {sheet['F4'].value}'
        assert sheet['G3'].value == 'Тип*', f'Expected Тип*, but got {sheet['G3'].value}'
        assert sheet['G4'].value == 'Заказчик', f'Expected Заказчик, but got {sheet['G4'].value}'
        assert sheet['H3'].value == 'Роль пользователя*', f'Expected Роль пользователя*, but got {sheet['H3'].value}'
        assert sheet['H4'].value == role.strip(), f'Expected {role.strip()}, but got {sheet['H4'].value}'
        assert sheet['I3'].value == 'Участок', f'Expected Участок, but got {sheet['I3'].value}'
        assert sheet['I4'].value == district_name.strip(), f'Expected {district_name.strip()}, but got {sheet['I4'].value}'
        assert sheet['J3'].value == 'Компания', f'Expected Компания, but got {sheet['J3'].value}'
        assert sheet['J4'].value == company_name.strip(), f'Expected {company_name.strip()}, but got {sheet['J4'].value}'

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Пользователи.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully export of list customers by userID.')

    @allure.step("Exports the list of users taking into account the specified filters (Staff).")
    def get_export_list_staff_by_user_id(
            self,
            user_id: int,
            name: str,
            surname: str,
            email: str,
            phone: str,
            role: str,
            district_name: str,
            technician: bool
    ):
        if technician is True:
            type_user = "Мобильный инженер"
        else:
            type_user = "Штатный сотрудник"
        params = {
            "userID": user_id,
            "isCustomer": "false",
            "isTeam": "false",
            "isBanned": "false",
            "isDeleted": "false",
            "isInitial": "false"
        }
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_list_users_endpoint, params=params,
            headers=self.headers.export_header(API_TOKEN)
        )
        end = time.time()
        logger.info(response.headers)
        self.attach_response_headers(response.headers)
        self.attach_time(start, end)
        self.attach_url(response.request.url)
        try:
            self.attach_response(response.json())
        except JSONDecodeError:
            logger.warning("Received response is not a valid JSON")
        assert response.status_code == HTTPStatus.OK, \
            f'Expected status code {HTTPStatus.OK}, but got {response.status_code}, {response.json()}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        # СОХРАНЕНИЕ ЛОКАЛЬНО EXCEL
        # output_file_path = "exported_data.xlsx"
        # workbook.save(output_file_path)

        sheet = workbook.active
        sheet_name = workbook.sheetnames

        # ВЫВОД СТРОК EXCEL ФАЙЛА
        # count = 0
        # for row in sheet.iter_rows(values_only=True):
        #     logger.warning(row)
        #     count += 1
        #     if count >= 10:
        #         break

        assert 'Пользователи' in sheet_name
        assert sheet['A3'].value == 'Фамилия*', f'Expected Фамилия*, but got {sheet['A3'].value}'
        assert sheet['A4'].value == surname.strip(), f'Expected {surname.strip()}, but got {sheet['A4'].value}'
        assert sheet['B3'].value == 'Имя*', f'Expected Имя*, but got {sheet['B3'].value}'
        assert sheet['B4'].value == name.strip(), f'Expected {name.strip()}, but got {sheet['B4'].value}'
        assert sheet['C3'].value == 'Отчество', f'Expected Отчество, but got {sheet['C3'].value}'
        assert sheet['D3'].value == 'Пол', f'Expected Пол, but got {sheet['D3'].value}'
        assert sheet['E3'].value == 'Телефон*', f'Expected Телефон*, but got {sheet['E3'].value}'
        assert f"+{sheet['E4'].value}" == phone.strip(), f'Expected {phone.strip()}, but got +{sheet['E4'].value}'
        assert sheet['F3'].value == 'Электронная почта*', f'Expected Электронная почта*, but got {sheet['F3'].value}'
        assert sheet['F4'].value == email.strip(), f'Expected {email.strip()}, but got {sheet['F4'].value}'
        assert sheet['G3'].value == 'Тип*', f'Expected Тип*, but got {sheet['G3'].value}'
        assert sheet['G4'].value == type_user, f'Expected {type_user}, but got {sheet['G4'].value}'
        assert sheet['H3'].value == 'Роль пользователя*', f'Expected Роль пользователя*, but got {sheet['H3'].value}'
        assert sheet['H4'].value == role.strip(), f'Expected {role.strip()}, but got {sheet['H4'].value}'
        assert sheet['I3'].value == 'Участок', f'Expected Участок, but got {sheet['I3'].value}'
        assert sheet['I4'].value == district_name.strip(), f'Expected {district_name.strip()}, but got {sheet['I4'].value}'
        assert sheet['J3'].value == 'Компания', f'Expected Компания, but got {sheet['J3'].value}'

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Пользователи.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully export of list staff by userID.')
