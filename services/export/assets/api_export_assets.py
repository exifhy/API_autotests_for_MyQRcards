from urllib import parse
import allure
import requests
from loguru import logger
from requests import JSONDecodeError
from utils.helper import Helper
from services.export.assets.payloads import Payloads
from services.export.assets.endpoints import Endpoints
from config.headers import Headers
from services.export.assets.models.export_assets_model import *
import time
from http import HTTPStatus
from dotenv import load_dotenv
import os
from faker import Faker
from openpyxl import load_workbook
from io import BytesIO


fake_ru = Faker('ru_RU')

load_dotenv()
API_TOKEN = os.getenv('API_TOKEN')


class ExportAssetsAPI(Helper):

    def __init__(self):
        super().__init__()
        self.payloads = Payloads()
        self.endpoints = Endpoints()
        self.headers = Headers()

    @allure.step("Returns a list of data available for extended exports.")
    def get_list_of_data_available_for_extended_exports(self):
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_list_object_extended_includes_endpoint,
            headers=self.headers.basic_header(API_TOKEN)
        )
        end = time.time()
        logger.info(response.headers)
        try:
            self.attach_response(response.json())
        except JSONDecodeError:
            logger.warning("Received response is not a valid JSON")
        self.attach_time(start, end)
        self.attach_url(response.request.url)
        assert response.status_code in {HTTPStatus.OK, HTTPStatus.PARTIAL_CONTENT}, f'{response.status_code}, {response.json()}'
        model = SuccessExportDataListModel(result=response.json())
        logger.info(f'Successfully receiving the list of data available for extended exports.')
        return model

    @allure.step("Normal export a list of objects.")
    def get_normal_export_list_objects(self):
        start = time.time()
        response = requests.get(
            url=self.endpoints.normal_export_list_object_endpoint,
            headers=self.headers.export_header(API_TOKEN)
        )
        end = time.time()
        logger.info(response.headers)
        self.attach_time(start, end)
        self.attach_url(response.request.url)
        try:
            logger.warning(response.json())
        except JSONDecodeError:
            logger.warning("Received response is not a valid JSON")
        assert response.status_code == HTTPStatus.OK, f'{response.status_code}, {response.json()}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        # СОХРАНЕНИЕ ЛОКАЛЬНО EXCEL
        # output_file_path = "exported_data.xlsx"
        # workbook.save(output_file_path)

        sheet = workbook.active
        sheet_name = workbook.sheetnames

        # ВЫВОД СТРОК EXCEL ФАЙЛА
        # count = 0
        # for row in sheet.iter_rows(values_only=True):
        #     logger.warning(row)
        #     count += 1
        #     if count >= 10:
        #         break

        assert 'Объекты и оборудование' in sheet_name
        assert sheet['A3'].value == 'Название*', \
            f"Expected <Название*>, but got <{sheet['A3'].value}>"

        assert sheet['B3'].value == 'ERP ID', \
            f"Expected <ERP ID>, but got <{sheet['B3'].value}>"

        assert sheet['C3'].value == 'Компания*', \
            f"Expected <Компания*>, but got <{sheet['C3'].value}>"

        assert sheet['D3'].value == 'Тип*', \
            f"Expected <Тип*>, but got <{sheet['D3'].value}>"

        assert sheet['E3'].value == 'Класс*', \
            f"Expected <Класс*>, but got <{sheet['E3'].value}>"

        assert sheet['F3'].value == 'Участок*', \
            f"Expected <Участок*>, but got <{sheet['F3'].value}>"

        assert sheet['G3'].value == 'Вид работ*', \
            f"Expected <Вид работ*>, but got <{sheet['G3'].value}>"

        assert sheet['I3'].value == 'Адрес*', \
            f"Expected <Адрес*>, but got <{sheet['I3'].value}>"

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Объекты+и+оборудование.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully receiving the normal export of list object.')

    @allure.step("Exports a list of objects with a set of filters by assetID(filter set in test case 23132).")
    def get_export_list_with_set_filter_by_asset_id(
            self,
            name_asset: str,
            asset_id: int,
            company_name: str,
            district_name: str,
            asset_type_name: str,
            asset_class_name: str,
            work_type_name: str
    ):
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_all_filters_by_asset_id_endpoint(asset_id),
            headers=self.headers.export_header(API_TOKEN)
        )
        end = time.time()
        logger.info(response.headers)
        self.attach_time(start, end)
        self.attach_url(response.request.url)
        try:
            self.attach_response(response.json())
        except JSONDecodeError:
            logger.warning("Received response is not a valid JSON")

        assert response.status_code == HTTPStatus.OK, f'{response.status_code}, {response.json()}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        # output_file_path = "exported_data.xlsx"
        # workbook.save(output_file_path)

        sheet = workbook.active
        sheet_name = workbook.sheetnames

        # count = 0
        # for row in sheet.iter_rows(values_only=True):
        #     logger.warning(row)
        #     count += 1
        #     if count >= 5:
        #         break

        # logger.info(sheet_name)
        assert 'Объекты и оборудование' in sheet_name

        assert sheet['A3'].value == 'Название', \
            f'Expected <Название>, but got <{sheet['A3'].value}>'

        assert sheet['A4'].value == name_asset.strip(), \
            f'Expected <{name_asset.strip()}>, but got <{sheet['A4'].value}>'

        assert sheet['B3'].value == 'Компания', \
            f'Expected <Компания>, but got <{sheet['B3'].value}>'

        assert sheet['B4'].value == company_name.strip(), \
            f'Expected <{company_name.strip()}>, but got <{sheet['B4'].value}>'

        assert sheet['C3'].value == 'Тип объекта', \
            f'Expected <Тип объекта>, but got <{sheet['C3'].value}>'

        assert sheet['C4'].value == asset_type_name.strip(), \
            f'Expected <{sheet['C4'].value}>, but got <{sheet['C4'].value}>'

        assert sheet['D3'].value == 'Класс объекта', \
            f'Expected <Класс объекта>, but got <{sheet['D3'].value}>'

        assert sheet['D4'].value == asset_class_name.strip(), \
            f'Expected <{sheet['D4'].value}>, but got <{sheet['D4'].value}>'

        assert sheet['E3'].value == 'Участок', \
            f'Expected <Участок>, but got <{sheet['E3'].value}>'

        assert sheet['E4'].value == district_name.strip(), \
            f'Expected <{district_name.strip()}>, but got <{sheet['E4'].value}>'

        assert sheet['F3'].value == 'Вид работ', \
            f'Expected <Вид работ>, but got <{sheet['F3'].value}>'

        assert sheet['F4'].value == work_type_name.strip(), \
            f'Expected <{work_type_name.strip()}>, but got <{sheet['F4'].value}>'

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Объекты+и+оборудование.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully receiving the xlsx with set filter by assetID.')
