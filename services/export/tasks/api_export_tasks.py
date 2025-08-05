import allure
import requests
from loguru import logger
from utils.helper import Helper
from services.export.tasks.payloads import Payloads
from services.export.tasks.endpoints import Endpoints
from config.headers import Headers
from services.export.tasks.models.export_tasks_model import *
import time
from http import HTTPStatus
from utils.token_utils import get_token
from openpyxl import load_workbook
from io import BytesIO
from urllib import parse


class ExportTasksAPI(Helper):

    def __init__(self):
        super().__init__()
        self.payloads = Payloads()
        self.endpoints = Endpoints()
        self.headers = Headers()

    @allure.step("Get list of tasks extended includes, check ScheduledFinishDateTime, ScheduledStartDateTime.")
    def get_list_of_tasks_extended_includes(self):
        start = time.time()
        response = requests.get(
            url=self.endpoints.get_list_tasks_extended_endpoint,
            headers=self.headers.export_header(get_token())
        )
        end = time.time()
        logger.info(response.headers)
        self.attach_response_headers(response.headers)
        data_response = self.response_content(response)
        self.attach_response(data_response)
        self.attach_time(start, end)
        self.attach_url(response.request.url)
        assert response.status_code in {HTTPStatus.OK, HTTPStatus.PARTIAL_CONTENT}, \
            (f'Expected status code {HTTPStatus.OK, HTTPStatus.PARTIAL_CONTENT}, '
             f'but got {response.status_code}, {data_response}')
        model = SuccessTasksResultModel(list=response.json())
        finish_item = next(item for item in model.list if item.code == "ScheduledFinishDateTime")
        start_item = next(item for item in model.list if item.code == "ScheduledStartDateTime")
        assert finish_item.description == "Назначено По", \
            f'Expected Назначено По, but got {finish_item.description}'
        assert start_item.description == "Назначено С", \
            f'Expected Назначено С, but got {start_item.description}'
        logger.info(f'Successfully get a list of data available for extended exports.')
        return model

    @allure.step("Exports the task list into account the specified filters by task id.")
    def get_normal_export_task_by_task_id(self, task_id: int, number_task: str, name_task_type: str):
        params = {
            "taskID": task_id,
            "isClosed": False,
            "isDeleted": False,
            "isInitial": False,
            "orderBy": 1
        }
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_list_tasks_endpoint, params=params,
            headers=self.headers.export_header(get_token())
        )
        end = time.time()
        logger.info(response.headers)
        self.attach_response_headers(response.headers)
        data_response = self.response_content(response)
        self.attach_time(start, end)
        self.attach_url(response.request.url)
        assert response.status_code == HTTPStatus.OK, \
            f'Expected status code {HTTPStatus.OK}, but got {response.status_code}, {data_response}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        # СОХРАНЕНИЕ ЛОКАЛЬНО EXCEL
        # output_file_path = "exported_data.xlsx"
        # workbook.save(output_file_path)

        sheet = workbook.active
        sheet_name = workbook.sheetnames

        assert 'Заявки' in sheet_name
        assert sheet['A3'].value == 'Номер', f'Expected Номер, but got {sheet['A3'].value}'
        assert sheet['A4'].value == number_task.strip(), f'Expected {number_task.strip()}, but got {sheet['A4'].value}'
        assert sheet['B3'].value == 'Тип заявки*', f'Expected Тип заявки*, but got {sheet['B3'].value}'
        assert sheet['B4'].value == name_task_type.strip(), \
            f'Expected <{name_task_type.strip()}>, but got {sheet['B4'].value}'
        assert sheet['C3'].value == 'Описание заявки', f'Expected Описание заявки, but got {sheet['C3'].value}'
        assert sheet['C4'].value == 'Заявка создана авто-тестом', \
            f'Expected <Заявка создана авто-тестом>, but got {sheet['C4'].value}'

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Заявки.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully export of list task by task id.')

    @allure.step("Exports the task list into account the specified filters by task number.")
    def get_normal_export_task_by_task_number(self, model_task, model_contact):
        params = {
            "isClosed": False,
            "isDeleted": False,
            "isInitial": False,
            "orderBy": 1,
            "searchText": model_task.number,
            "sortDirection": 2
        }
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_list_tasks_endpoint, params=params,
            headers=self.headers.export_header(get_token())
        )
        end = time.time()
        logger.info(response.headers)
        self.attach_response_headers(response.headers)
        data_response = self.response_content(response)
        self.attach_time(start, end)
        self.attach_url(response.request.url)
        assert response.status_code == HTTPStatus.OK, \
            f'Expected status code {HTTPStatus.OK}, but got {response.status_code}, {data_response}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        sheet = workbook.worksheets[0]
        sheet_names = workbook.sheetnames

        assert 'Заявки' == sheet_names[0], f"Expected Заявки, but got {sheet_names[0]}"

        assert sheet['A3'].value == 'Номер', f'Expected Номер, but got {sheet["A3"].value}'
        assert sheet['A4'].value == model_task.number, f'Expected <{model_task.number}>, but got {sheet["A4"].value}'
        assert sheet['B3'].value == 'Тип заявки*', f'Expected Тип заявки*, but got {sheet["B3"].value}'
        assert sheet['B4'].value == model_task.taskType.name.strip(), \
            f'Expected <{model_task.taskType.name.strip()}>, but got {sheet["B4"].value}'

        assert sheet['C3'].value == 'Описание заявки', f'Expected Описание заявки, but got {sheet["C3"].value}'
        assert sheet['C4'].value == model_task.notes.strip(), \
            f'Expected <{model_task.notes.strip()}>, but got {sheet["C4"].value}'
        assert sheet['D3'].value == 'Оборудование*', f'Expected Оборудование*, but got {sheet["D3"].value}'
        assert sheet['D4'].value == model_task.asset.name.strip(), \
            f'Expected <{model_task.asset.name.strip()}>, but got {sheet["D4"].value}'
        assert sheet['E3'].value == 'Вид работ*', f'Expected Вид работ*, but got {sheet["E3"].value}'
        assert sheet['E4'].value == model_task.workType.name.strip(), \
            f'Expected <{model_task.workType.name.strip()}>, but got {sheet["E4"].value}'
        assert sheet['F3'].value == 'Крайний срок закрытия', \
            f'Expected Крайний срок закрытия, but got {sheet["F3"].value}'

        assert sheet['G3'].value == 'Критичность', f'Expected Критичность, but got {sheet["G3"].value}'
        assert sheet['G4'].value == model_task.actualCriticality.name, \
            f'Expected <{model_task.actualCriticality.name}>, but got {sheet["G4"].value}'
        assert sheet['H3'].value == 'Исполнитель', f'Expected Исполнитель, but got {sheet["H3"].value}'

        assigned_to = ",".join(
            f"{stuff.lastName} {stuff.firstName} {stuff.middleName if stuff.middleName else ''}".strip()
            for stuff in model_task.listAssignedTo
        )

        assert sheet['H4'].value == assigned_to, f'Expected <{assigned_to}>, but got {sheet["H4"].value}'
        assert sheet['I3'].value == 'Дата назначения', f'Expected Дата назначения, but got {sheet["I3"].value}'
        # assert sheet['I4'].value == '', f'Expected <>, but got {sheet["I4"].value}'
        assert sheet['J3'].value == 'Назначено С', f'Expected Назначено С, but got {sheet["J3"].value}'
        # assert sheet['J4'].value == '', f'Expected <>, but got {sheet["J4"].value}'
        assert sheet['K3'].value == 'Назначено По', f'Expected Назначено По, but got {sheet["K3"].value}'
        # assert sheet['K4'].value == '', f'Expected <>, but got {sheet["K4"].value}'
        assert sheet['L3'].value == 'Заказчик', f'Expected Заказчик, but got {sheet["L3"].value}'
        assert sheet['L4'].value == model_task.company.name.strip(), \
            f'Expected <{model_task.company.name.strip()}>, but got {sheet["L4"].value}'
        assert sheet['M3'].value == 'Адрес', f'Expected Адрес, but got {sheet["M3"].value}'
        assert sheet['M4'].value == model_task.location.address, \
            f'Expected <{model_task.location.address}>, but got {sheet["M4"].value}'
        assert sheet['N3'].value == 'Описание объекта', f'Expected Описание объекта, but got {sheet["N3"].value}'
        assert sheet['N4'].value == model_task.location.description.strip(), \
            f'Expected <{model_task.location.description.strip()}>, but got {sheet["N4"].value}'

        latitude, longitude = model_task.location.coordinate.split(":")

        latitude, longitude = float(latitude), float(longitude)

        assert sheet['O3'].value == 'Широта', f'Expected Широта, but got {sheet["O3"].value}'
        assert sheet['O4'].value == latitude, f'Expected <{latitude}>, but got {sheet["O4"].value}'
        assert sheet['P3'].value == 'Долгота', f'Expected Долгота, but got {sheet["P3"].value}'
        assert sheet['P4'].value == longitude, f'Expected <{longitude}>, but got {sheet["P4"].value}'

        assert sheet['Q3'].value == 'Часовой пояс', f'Expected Часовой пояс, but got {sheet["Q3"].value}'
        assert sheet['Q4'].value == model_task.location.timeZone.name.strip(), \
            f'Expected <{model_task.location.timeZone.name.strip()}>, but got {sheet["Q4"].value}'

        assert sheet['R3'].value == 'Страна', f'Expected Страна, but got {sheet["R3"].value}'
        assert sheet['R4'].value.strip() == model_task.location.country.name.strip(), \
            f'Expected <{model_task.location.country.name}>, but got {sheet["R4"].value}'

        assert sheet['S3'].value == 'Стадия', f'Expected Стадия, but got {sheet["S3"].value}'
        assert sheet['S4'].value == model_task.taskStage.name, \
            f'Expected <{model_task.taskStage.name}>, but got {sheet["S4"].value}'
        assert sheet['T3'].value == 'Статус', f'Expected Статус, but got {sheet["T3"].value}'
        assert sheet['T4'].value == model_task.taskStatus.name, \
            f'Expected <{model_task.taskStatus.name}>, but got {sheet["T4"].value}'
        assert sheet['U3'].value == 'Актуальность', f'Expected Актуальность, but got {sheet["U3"].value}'
        # assert sheet['U4'].value == '', f'Expected <>, but got {sheet["U4"].value}'
        assert sheet['V3'].value == 'Контактное лицо', f'Expected Контактное лицо, but got {sheet["V3"].value}'
        assert sheet['V4'].value == model_contact.fullName.strip(), \
            f'Expected <{model_contact.fullName.strip()}>, but got {sheet["V4"].value}'
        assert sheet['W3'].value == 'Электронная почта', f'Expected Электронная почта, but got {sheet["W3"].value}'
        assert sheet['W4'].value == model_contact.email.strip(), \
            f'Expected <{model_contact.email.strip()}>, but got {sheet["W4"].value}'
        assert sheet['X3'].value == 'Телефон', f'Expected Телефон, but got {sheet["X3"].value}'
        assert sheet['X4'].value == model_contact.phone.strip(), \
            f'Expected <{model_contact.phone.strip()}>, but got {sheet["X4"].value}'
        assert sheet['Y3'].value == 'Обращение от', f'Expected Обращение от, but got {sheet["Y3"].value}'
        assert sheet['Y4'].value == model_task.requestedBy.firstName.strip(), \
            f'Expected <{model_task.requestedBy.firstName.strip()}>, but got {sheet["Y4"].value}'

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Заявки.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully normal export of list task by task number.')

    @allure.step("Exports the task list into account the specified filters by task number, V2.")
    def get_normal_export_task_by_task_number_v2(self, model_task, model_contact):
        params = {
            "isClosed": False,
            "isDeleted": False,
            "isInitial": False,
            "orderBy": 1,
            "searchText": model_task.number,
            "sortDirection": 2
        }
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_list_tasks_v2_endpoint, params=params,
            headers=self.headers.export_header(get_token())
        )
        end = time.time()
        logger.info(response.headers)
        self.attach_response_headers(response.headers)
        data_response = self.response_content(response)
        self.attach_time(start, end)
        self.attach_url(response.request.url)
        assert response.status_code == HTTPStatus.OK, \
            f'Expected status code {HTTPStatus.OK}, but got {response.status_code}, {data_response}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        sheet = workbook.worksheets[0]
        sheet_names = workbook.sheetnames

        assert 'Заявки' == sheet_names[0], f"Expected Заявки, but got {sheet_names[0]}"

        assert sheet['A3'].value == 'Номер', f'Expected Номер, but got {sheet["A3"].value}'
        assert sheet['A4'].value == model_task.number, f'Expected <{model_task.number}>, but got {sheet["A4"].value}'
        assert sheet['B3'].value == 'Тип заявки*', f'Expected Тип заявки*, but got {sheet["B3"].value}'
        assert sheet['B4'].value == model_task.taskType.name.strip(), \
            f'Expected <{model_task.taskType.name.strip()}>, but got {sheet["B4"].value}'

        assert sheet['C3'].value == 'Описание заявки', f'Expected Описание заявки, but got {sheet["C3"].value}'
        assert sheet['C4'].value == model_task.notes.strip(), \
            f'Expected <{model_task.notes.strip()}>, but got {sheet["C4"].value}'
        assert sheet['D3'].value == 'Оборудование*', f'Expected Оборудование*, but got {sheet["D3"].value}'
        assert sheet['D4'].value == model_task.asset.name.strip(), \
            f'Expected <{model_task.asset.name.strip()}>, but got {sheet["D4"].value}'
        assert sheet['E3'].value == 'Вид работ*', f'Expected Вид работ*, but got {sheet["E3"].value}'
        assert sheet['E4'].value == model_task.workType.name.strip(), \
            f'Expected <{model_task.workType.name.strip()}>, but got {sheet["E4"].value}'
        assert sheet['F3'].value == 'Крайний срок закрытия', \
            f'Expected Крайний срок закрытия, but got {sheet["F3"].value}'

        assert sheet['G3'].value == 'Критичность', f'Expected Критичность, but got {sheet["G3"].value}'
        assert sheet['G4'].value == model_task.actualCriticality.name, \
            f'Expected <{model_task.actualCriticality.name}>, but got {sheet["G4"].value}'
        assert sheet['H3'].value == 'Исполнитель', f'Expected Исполнитель, but got {sheet["H3"].value}'

        assigned_to = ",".join(
            f"{stuff.lastName} {stuff.firstName} {stuff.middleName if stuff.middleName else ''}".strip()
            for stuff in model_task.listAssignedTo
        )

        assert sheet['H4'].value == assigned_to, f'Expected <{assigned_to}>, but got {sheet["H4"].value}'
        assert sheet['I3'].value == 'Дата назначения', f'Expected Дата назначения, but got {sheet["I3"].value}'
        # assert sheet['I4'].value == '', f'Expected <>, but got {sheet["I4"].value}'
        assert sheet['J3'].value == 'Назначено С', f'Expected Назначено С, but got {sheet["J3"].value}'
        # assert sheet['J4'].value == '', f'Expected <>, but got {sheet["J4"].value}'
        assert sheet['K3'].value == 'Назначено По', f'Expected Назначено По, but got {sheet["K3"].value}'
        # assert sheet['K4'].value == '', f'Expected <>, but got {sheet["K4"].value}'
        assert sheet['L3'].value == 'Заказчик', f'Expected Заказчик, but got {sheet["L3"].value}'
        assert sheet['L4'].value == model_task.company.name.strip(), \
            f'Expected <{model_task.company.name.strip()}>, but got {sheet["L4"].value}'
        assert sheet['M3'].value == 'Адрес', f'Expected Адрес, but got {sheet["M3"].value}'
        assert sheet['M4'].value == model_task.location.address, \
            f'Expected <{model_task.location.address}>, but got {sheet["M4"].value}'
        assert sheet['N3'].value == 'Описание объекта', f'Expected Описание объекта, but got {sheet["N3"].value}'
        assert sheet['N4'].value == model_task.location.description.strip(), \
            f'Expected <{model_task.location.description.strip()}>, but got {sheet["N4"].value}'

        latitude, longitude = model_task.location.coordinate.split(":")

        latitude, longitude = float(latitude), float(longitude)

        assert sheet['O3'].value == 'Широта', f'Expected Широта, but got {sheet["O3"].value}'
        assert sheet['O4'].value == latitude, f'Expected <{latitude}>, but got {sheet["O4"].value}'
        assert sheet['P3'].value == 'Долгота', f'Expected Долгота, but got {sheet["P3"].value}'
        assert sheet['P4'].value == longitude, f'Expected <{longitude}>, but got {sheet["P4"].value}'

        assert sheet['Q3'].value == 'Часовой пояс', f'Expected Часовой пояс, but got {sheet["Q3"].value}'
        assert sheet['Q4'].value == model_task.location.timeZone.name.strip(), \
            f'Expected <{model_task.location.timeZone.name.strip()}>, but got {sheet["Q4"].value}'

        assert sheet['R3'].value == 'Страна', f'Expected Страна, but got {sheet["R3"].value}'
        assert sheet['R4'].value.strip() == model_task.location.country.name.strip(), \
            f'Expected <{model_task.location.country.name}>, but got {sheet["R4"].value}'

        assert sheet['S3'].value == 'Стадия', f'Expected Стадия, but got {sheet["S3"].value}'
        assert sheet['S4'].value == model_task.taskStage.name, \
            f'Expected <{model_task.taskStage.name}>, but got {sheet["S4"].value}'
        assert sheet['T3'].value == 'Статус', f'Expected Статус, but got {sheet["T3"].value}'
        assert sheet['T4'].value == model_task.taskStatus.name, \
            f'Expected <{model_task.taskStatus.name}>, but got {sheet["T4"].value}'
        assert sheet['U3'].value == 'Актуальность', f'Expected Актуальность, but got {sheet["U3"].value}'
        # assert sheet['U4'].value == '', f'Expected <>, but got {sheet["U4"].value}'
        assert sheet['V3'].value == 'Контактное лицо', f'Expected Контактное лицо, but got {sheet["V3"].value}'
        assert sheet['V4'].value == model_contact.fullName.strip(), \
            f'Expected <{model_contact.fullName.strip()}>, but got {sheet["V4"].value}'
        assert sheet['W3'].value == 'Электронная почта', f'Expected Электронная почта, but got {sheet["W3"].value}'
        assert sheet['W4'].value == model_contact.email.strip(), \
            f'Expected <{model_contact.email.strip()}>, but got {sheet["W4"].value}'
        assert sheet['X3'].value == 'Телефон', f'Expected Телефон, but got {sheet["X3"].value}'
        assert sheet['X4'].value == model_contact.phone.strip(), \
            f'Expected <{model_contact.phone.strip()}>, but got {sheet["X4"].value}'
        assert sheet['Y3'].value == 'Обращение от', f'Expected Обращение от, but got {sheet["Y3"].value}'
        assert sheet['Y4'].value == model_task.requestedBy.firstName.strip(), \
            f'Expected <{model_task.requestedBy.firstName.strip()}>, but got {sheet["Y4"].value}'

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Заявки.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully normal export of list task by task number, V2.')

    @allure.step("Export an empty template for importing task.")
    def get_export_empty_template_for_importing_task(self):
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_empty_template_for_importing_task_endpoint,
            headers=self.headers.export_header(get_token())
        )
        end = time.time()
        logger.info(response.headers)
        self.attach_response_headers(response.headers)
        data_response = self.response_content(response)
        self.attach_time(start, end)
        self.attach_url(response.request.url)
        assert response.status_code == HTTPStatus.OK, \
            f'Expected status code {HTTPStatus.OK}, but got {response.status_code}, {data_response}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        sheet = workbook.worksheets[0]
        sheet_names = workbook.sheetnames

        assert 'Заявки' == sheet_names[0], f"Expected Заявки, but got {sheet_names[0]}"

        assert sheet['A3'].value == 'Номер', f'Expected Номер, but got {sheet["A3"].value}'
        assert sheet['B3'].value == 'Тип заявки*', f'Expected Тип заявки*, but got {sheet["B3"].value}'
        assert sheet['C3'].value == 'Описание заявки', f'Expected Описание заявки, but got {sheet["C3"].value}'
        assert sheet['D3'].value == 'Оборудование*', f'Expected Оборудование*, but got {sheet["D3"].value}'
        assert sheet['E3'].value == 'Вид работ*', f'Expected Вид работ*, but got {sheet["E3"].value}'
        assert sheet['F3'].value == 'Крайний срок закрытия', \
            f'Expected Крайний срок закрытия, but got {sheet["F3"].value}'
        assert sheet['G3'].value == 'Критичность', f'Expected Критичность, but got {sheet["G3"].value}'
        assert sheet['H3'].value == 'Исполнитель', f'Expected Исполнитель, but got {sheet["H3"].value}'
        assert sheet['I3'].value == 'Дата назначения', f'Expected Дата назначения, but got {sheet["I3"].value}'
        assert sheet['J3'].value == 'Назначено С', f'Expected Назначено С, but got {sheet["J3"].value}'
        assert sheet['K3'].value == 'Назначено По', f'Expected Назначено По, but got {sheet["K3"].value}'
        assert sheet['L3'].value == 'Заказчик', f'Expected Заказчик, but got {sheet["L3"].value}'
        assert sheet['M3'].value == 'Адрес', f'Expected Адрес, but got {sheet["M3"].value}'
        assert sheet['M2'].value == 'Адрес объекта', f'Expected Адрес объекта, but got {sheet["M2"].value}'
        assert sheet['N3'].value == 'Описание объекта', f'Expected Описание объекта, but got {sheet["N3"].value}'
        assert sheet['O3'].value == 'Широта', f'Expected Широта, but got {sheet["O3"].value}'
        assert sheet['P3'].value == 'Долгота', f'Expected Долгота, but got {sheet["P3"].value}'
        assert sheet['Q3'].value == 'Часовой пояс', f'Expected Часовой пояс, but got {sheet["Q3"].value}'
        assert sheet['R3'].value == 'Страна', f'Expected Страна, but got {sheet["R3"].value}'
        assert sheet['S3'].value == 'Стадия', f'Expected Стадия, but got {sheet["S3"].value}'
        assert sheet['T3'].value == 'Статус', f'Expected Статус, but got {sheet["T3"].value}'
        assert sheet['U3'].value == 'Актуальность', f'Expected Актуальность, but got {sheet["U3"].value}'
        assert sheet['V3'].value == 'Контактное лицо', f'Expected Контактное лицо, but got {sheet["V3"].value}'
        assert sheet['W3'].value == 'Электронная почта', f'Expected Электронная почта, but got {sheet["W3"].value}'
        assert sheet['X3'].value == 'Телефон', f'Expected Телефон, but got {sheet["X3"].value}'
        assert sheet['Y3'].value == 'Обращение от', f'Expected Обращение от, but got {sheet["Y3"].value}'

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Заявки.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully export an empty template for importing task.')

    @allure.step("Exports the extended task list taking into account the specified filters by task number.")
    def get_extended_export_task_by_task_number(
            self, model_task, model_contact, model_asset, model_district
    ):
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_list_extended_tasks_all_filters_by_task_id_endpoint(model_task.number),
            headers=self.headers.export_header(get_token())
        )
        end = time.time()
        logger.info(response.headers)
        self.attach_response_headers(response.headers)
        data_response = self.response_content(response)
        self.attach_time(start, end)
        self.attach_url(response.request.url)
        assert response.status_code == HTTPStatus.OK, \
            f'Expected status code {HTTPStatus.OK}, but got {response.status_code}, {data_response}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        sheet = workbook.worksheets[0]
        sheet_second = workbook.worksheets[1]
        sheet_third = workbook.worksheets[2]
        sheet_fourth = workbook.worksheets[3]
        sheet_fifth = workbook.worksheets[4]
        sheet_names = workbook.sheetnames

        assert 'Заявки' == sheet_names[0], f"Expected Заявки, but got {sheet_names[0]}"
        assert sheet['A2'].value == 'Номер', f'Expected Номер, but got {sheet['A2'].value}'
        assert sheet['A3'].value == model_task.number, f'Expected {model_task.number}, but got {sheet['A3'].value}'
        assert sheet['B2'].value == 'Родительская заявка', f'Expected Родительская заявка, but got {sheet['B2'].value}'
        assert sheet['B3'].value == getattr(getattr(model_task, 'parent', None), 'name', ''), \
            f'Expected {getattr(getattr(model_task, 'parent', None), 'name', '')}, but got {sheet['B3'].value}'
        assert sheet['C2'].value == 'Тип заявки', f'Expected Тип заявки, but got {sheet['C2'].value}'
        assert sheet['C3'].value == model_task.taskType.name.strip(), \
            f'Expected {model_task.taskType.name.strip()}, but got {sheet['C3'].value}.'
        assert sheet['D2'].value == 'Описание заявки', f'Expected Описание заявки, but got {sheet['D2'].value}'
        assert sheet['D3'].value == model_task.notes.strip(), \
            f'Expected {model_task.notes.strip()}, but got {sheet['D3'].value}'

        assert sheet['E2'].value == 'Заказчик', f'Expected Заказчик, but got {sheet['E2'].value}'
        assert sheet['E3'].value == model_task.company.name.strip(), \
            f'Expected {model_task.company.name.strip()}, but got {sheet['E3'].value}'
        assert sheet['F2'].value == 'Оборудование', f'Expected Оборудование, but got {sheet['F2'].value}'
        assert sheet['F3'].value == model_task.asset.name.strip(), \
            f'Expected {model_task.asset.name.strip()}, but got {sheet['F3'].value}'
        assert sheet['G2'].value == 'Серийный номер объекта', \
            f'Expected Серийный номер объекта, but got {sheet['G2'].value}'
        assert sheet['H2'].value == 'Адрес', f'Expected Адрес, but got {sheet['H2'].value}'
        assert sheet['H3'].value == model_task.location.address, \
            f'Expected {model_task.location.address}, but got {sheet['H3'].value}'
        assert sheet['I2'].value == 'Описание объекта', f'Expected Описание объекта, but got {sheet['I2'].value}'
        assert sheet['I3'].value == model_task.location.description.strip(), \
            f'Expected {model_task.location.description.strip()}, but got {sheet['I3'].value}'

        latitude, longitude = model_task.location.coordinate.split(":")

        latitude, longitude = float(latitude), float(longitude)

        assert sheet['J2'].value == 'Широта', f'Expected Широта, but got {sheet['J2'].value}'
        assert sheet['J3'].value == latitude, f'Expected {latitude}, but got {sheet['J3'].value}'
        assert sheet['K2'].value == 'Долгота', f'Expected Долгота, but got {sheet['K2'].value}'
        assert sheet['K3'].value == longitude, f'Expected {longitude}, but got {sheet['K3'].value}'
        assert sheet['L2'].value == 'Часовой пояс', f'Expected Часовой пояс, but got {sheet['L2'].value}'
        assert sheet['L3'].value == model_task.location.timeZone.name.strip(), \
            f'Expected {model_task.location.timeZone.name.strip()}, but got {sheet['L3'].value}'
        assert sheet['M2'].value == 'Страна', f'Expected Страна, but got {sheet["M2"].value}'
        assert sheet['M3'].value == model_task.location.country.name, \
            f'Expected {model_task.location.country.name}, but got {sheet["M3"].value}'
        assert sheet['N2'].value == 'Вид работ', f'Expected Вид работ, but got {sheet["N2"].value}'
        assert sheet['N3'].value == model_task.workType.name.strip(), \
            f'Expected {model_task.workType.name.strip()}, but got {sheet["N3"].value}'

        assert sheet['O2'].value == 'Стадия', f'Expected Стадия, but got {sheet["O2"].value}'
        assert sheet['O3'].value == model_task.taskStage.name, \
            f'Expected {model_task.taskStage.name}, but got {sheet["O3"].value}'
        assert sheet['P2'].value == 'Статус', f'Expected Статус, but got {sheet["P2"].value}'
        assert sheet['P3'].value == model_task.taskStatus.name, \
            f'Expected {model_task.taskStatus.name}, but got {sheet["P3"].value}'

        assert sheet['Q2'].value == 'Актуальность', f'Expected Актуальность, but got {sheet["Q2"].value}'
        # assert sheet['Q3'].value == '', f'Expected , but got {sheet["Q3"].value}'
        assert sheet['R2'].value == 'Критичность', f'Expected Критичность, but got {sheet["R2"].value}'
        assert sheet['R3'].value == model_task.actualCriticality.name, \
            f'Expected {model_task.actualCriticality.name}, but got {sheet["R3"].value}'
        assert sheet['S2'].value == 'Крайний срок закрытия', \
            f'Expected Крайний срок закрытия, but got {sheet["S2"].value}'
        # assert sheet['S3'].value == '', f'Expected , but got {sheet["S3"].value}'
        assert sheet['T2'].value == 'Исполнитель', f'Expected Исполнитель, but got {sheet["T2"].value}'
        assigned_to = ",".join(
            f"{stuff.lastName} {stuff.firstName} {stuff.middleName if stuff.middleName else ''}".strip()
            for stuff in model_task.listAssignedTo
        )
        assert sheet['T3'].value == assigned_to, f'Expected {assigned_to}, but got {sheet["T3"].value}'

        assert sheet['U2'].value == 'Дата назначения', f'Expected Дата назначения, but got {sheet["U2"].value}'
        # assert sheet['U3'].value == '', f'Expected , but got {sheet["U3"].value}'
        assert sheet['V2'].value == 'Назначено С', f'Expected Назначено С, but got {sheet["V2"].value}'
        # assert sheet['V3'].value == '', f'Expected , but got {sheet["V3"].value}'
        assert sheet['W2'].value == 'Назначено По', f'Expected Назначено По, but got {sheet["W2"].value}'
        # assert sheet['W3'].value == '', f'Expected , but got {sheet["W3"].value}'

        assert sheet['X2'].value == 'Контактное лицо', f'Expected Контактное лицо, but got {sheet["X2"].value}'
        assert sheet['X3'].value == model_contact.fullName.strip(), \
            f'Expected {model_contact.fullName.strip()}, but got {sheet["X3"].value}'

        assert sheet['Y2'].value == 'Электронная почта', f'Expected Электронная почта, but got {sheet["Y2"].value}'
        assert sheet['Y3'].value == model_contact.email.strip(), \
            f'Expected {model_contact.email.strip()}, but got {sheet["Y3"].value}'

        assert sheet['Z2'].value == 'Телефон', f'Expected Телефон, but got {sheet["Z2"].value}'
        assert sheet['Z3'].value == model_contact.phone.strip(), \
            f'Expected {model_contact.phone.strip()}, but got {sheet["Z3"].value}'

        assert sheet['AA2'].value == 'Обращение от', f'Expected Обращение от, but got {sheet["AA2"].value}'
        assert sheet['AA3'].value == model_task.requestedBy.firstName.strip(), \
            f'Expected {model_task.requestedBy.firstName.strip()}, but got {sheet["AA3"].value}'

        assert sheet['AB2'].value == 'Ответственный за объект', \
            f'Expected Ответственный за объект, but got {sheet["AB2"].value}'
        responsible_person = (f"{model_asset.responsiblePerson.lastName} "
                              f"{model_asset.responsiblePerson.firstName} "
                              f"{model_asset.responsiblePerson.middleName if model_asset.responsiblePerson.middleName else ''}".strip()
                              )
        assert sheet['AB3'].value == responsible_person, \
            f'Expected {responsible_person}, but got {sheet["AB3"].value}'

        assert sheet['AC2'].value == 'Название участка', f'Expected Название участка, but got {sheet["AC2"].value}'

        district = model_district.root[next(iter(model_district.root))].name.strip()
        assert sheet['AC3'].value == district, f'Expected {district}, but got {sheet["AC3"].value}'

        assert sheet['AD2'].value == 'Класс объекта', f'Expected Класс объекта, but got {sheet["AD2"].value}'
        assert sheet['AD3'].value == model_asset.assetClass.name.strip(), \
            f'Expected {model_asset.assetClass.name.strip()}, but got {sheet["AD3"].value}'

        assert sheet['AE2'].value == 'Тип объекта', f'Expected Тип объекта, but got {sheet["AE2"].value}'
        assert sheet['AE3'].value == model_asset.assetType.name.strip(), \
            f'Expected {model_asset.assetType.name.strip()}, but got {sheet["AE3"].value}'

        assert sheet['AF2'].value == 'Неисправность обнаружена', \
            f'Expected Неисправность обнаружена, but got {sheet["AF2"].value}'
        # assert sheet['AF3'].value == '', f'Expected , but got {sheet["AF3"].value}'

        assert sheet['AG2'].value == 'Выполнена', f'Expected Выполнена, but got {sheet["AG2"].value}'
        # assert sheet['AG3'].value == '', f'Expected , but got {sheet["AG3"].value}'

        assert sheet['AH2'].value == 'Дата создания', f'Expected Дата создания, but got {sheet["AH2"].value}'
        # assert sheet['AH3'].value == '', f'Expected , but got {sheet["AH3"].value}'

        assert sheet['AI2'].value == 'Закрыта', f'Expected Закрыта, but got {sheet["AI2"].value}'
        # assert sheet['AI3'].value == '', f'Expected , but got {sheet["AI3"].value}'

        assert sheet['AJ2'].value == 'Объект посещен', f'Expected Объект посещен, but got {sheet["AJ2"].value}'
        # assert sheet['AJ3'].value == '', f'Expected , but got {sheet["AJ3"].value}'

        assert sheet['AK2'].value == 'Метод подачи заявки', \
            f'Expected Метод подачи заявки, but got {sheet["AK2"].value}'
        assert sheet['AK3'].value == model_task.requestMethod.name.strip(), \
            f'Expected {model_task.requestMethod.name.strip()}, but got {sheet["AK3"].value}'

        assert sheet['AL2'].value == 'Оценочные трудозатраты', \
            f'Expected Оценочные трудозатраты, but got {sheet["AL2"].value}'
        # assert sheet['AL3'].value == '', f'Expected , but got {sheet["AL3"].value}'

        assert sheet['AM2'].value == 'Оценочная стоимость', \
            f'Expected Оценочная стоимость, but got {sheet["AM2"].value}'
        # assert sheet['AM3'].value == '', f'Expected , but got {sheet["AM3"].value}'

        assert sheet['AN2'].value == 'Фактические трудозатраты', \
            f'Expected Фактические трудозатраты, but got {sheet["AN2"].value}'
        # assert sheet['AN3'].value == '', f'Expected , but got {sheet["AN3"].value}'

        assert sheet['AO2'].value == 'Фактическая стоимость', \
            f'Expected Фактическая стоимость, but got {sheet["AO2"].value}'
        # assert sheet['AO3'].value == '', f'Expected , but got {sheet["AO3"].value}'

        assert sheet['AP2'].value.strip() == 'Шаблоны', f'Expected Шаблоны, but got {sheet["AP2"].value}'
        # assert sheet['AP3'].value == '', f'Expected , but got {sheet["AP3"].value}'

        assert 'Выполненные работы' == sheet_names[1], f"Expected Выполненные работы, but got {sheet_names[1]}"
        assert sheet_second['A2'].value == 'Номер', f'Expected Номер, but got {sheet_second["A2"].value}'
        # assert sheet_second['A3'].value == '', f'Expected , but got {sheet_second["A3"].value}'
        assert sheet_second['B2'].value == 'Вид работ', f'Expected Вид работ, but got {sheet_second["B2"].value}'
        # assert sheet_second['B3'].value == '', f'Expected , but got {sheet_second["B3"].value}'
        assert sheet_second['C2'].value == 'Стоимость вида работ', \
            f'Expected Стоимость вида работ, but got {sheet_second["C2"].value}'
        # assert sheet_second['C3'].value == '', f'Expected , but got {sheet_second["C3"].value}'
        assert sheet_second['D2'].value == 'Оборудование', f'Expected Оборудование, but got {sheet_second["D2"].value}'
        # assert sheet_second['D3'].value == '', f'Expected , but got {sheet_second["D3"].value}'
        assert sheet_second['E2'].value == 'Выполнил', f'Expected Выполнил, but got {sheet_second["E2"].value}'
        # assert sheet_second['E3'].value == '', f'Expected , but got {sheet_second["E3"].value}'
        assert sheet_second['F2'].value == 'Ставка сотрудника', \
            f'Expected Ставка сотрудника, but got {sheet_second["F2"].value}'
        # assert sheet_second['F3'].value == '', f'Expected , but got {sheet_second["F3"].value}'
        assert sheet_second['G2'].value == 'Начало работ', f'Expected Начало работ, but got {sheet_second["G2"].value}'
        # assert sheet_second['G3'].value == '', f'Expected , but got {sheet_second["G3"].value}'
        assert sheet_second['H2'].value == 'Окончание работ', \
            f'Expected Окончание работ, but got {sheet_second["H2"].value}'
        # assert sheet_second['H3'].value == '', f'Expected , but got {sheet_second["H3"].value}'
        assert sheet_second['I2'].value == 'Объем работ', f'Expected Объем работ, but got {sheet_second["I2"].value}'
        # assert sheet_second['I3'].value == '', f'Expected , but got {sheet_second["I3"].value}'
        assert sheet_second['J2'].value == 'Единица измерения', \
            f'Expected Единица измерения, but got {sheet_second["J2"].value}'
        # assert sheet_second['J3'].value == '', f'Expected , but got {sheet_second["J3"].value}'
        assert sheet_second['K2'].value == 'Описание', f'Expected Описание, but got {sheet_second["K2"].value}'
        # assert sheet_second['K3'].value == '', f'Expected , but got {sheet_second["K3"].value}'

        assert 'Сообщения' == sheet_names[2], f"Expected Сообщения, but got {sheet_names[2]}"
        assert sheet_third['A2'].value == 'Номер', f'Expected Номер, but got {sheet_third["A2"].value}'
        # assert sheet_third['A3'].value == '', f'Expected , but got {sheet_third["A3"].value}'
        assert sheet_third['B2'].value == 'Оправитель', f'Expected Оправитель, but got {sheet_third["B2"].value}'
        # assert sheet_third['B3'].value == '', f'Expected , but got {sheet_third["B3"].value}'
        assert sheet_third['C2'].value == 'Сообщение', \
            f'Expected Сообщение, but got {sheet_third["C2"].value}'
        # assert sheet_third['C3'].value == '', f'Expected , but got {sheet_third["C3"].value}'
        assert sheet_third['D2'].value == 'Отправлено', f'Expected Отправлено, but got {sheet_third["D2"].value}'
        # assert sheet_third['D3'].value == '', f'Expected , but got {sheet_third["D3"].value}'

        assert 'Чек листы' == sheet_names[3], f"Expected Чек листы, but got {sheet_names[3]}"
        assert sheet_fourth['A2'].value == 'Номер', f'Expected Номер, but got {sheet_fourth["A2"].value}'
        # assert sheet_fourth['A3'].value == '', f'Expected , but got {sheet_fourth["A3"].value}'
        assert sheet_fourth['B2'].value == 'Чек лист', f'Expected Чек лист, but got {sheet_fourth["B2"].value}'
        # assert sheet_fourth['B3'].value == '', f'Expected , but got {sheet_fourth["B3"].value}'
        assert sheet_fourth['C2'].value == 'Пункт чек листа', \
            f'Expected Пункт чек листа, but got {sheet_fourth["C2"].value}'
        # assert sheet_fourth['C3'].value == '', f'Expected , but got {sheet_fourth["C3"].value}'
        assert sheet_fourth['D2'].value == 'Отмечено', f'Expected Отмечено, but got {sheet_fourth["D2"].value}'
        # assert sheet_fourth['D3'].value == '', f'Expected , but got {sheet_fourth["D3"].value}'
        assert sheet_fourth['E2'].value == 'Значение', f'Expected Значение, but got {sheet_fourth["E2"].value}'
        # assert sheet_fourth['E3'].value == '', f'Expected , but got {sheet_fourth["E3"].value}'
        assert sheet_fourth['F2'].value == 'Обработано', \
            f'Expected Обработано, but got {sheet_fourth["F2"].value}'
        # assert sheet_fourth['F3'].value == '', f'Expected , but got {sheet_fourth["F3"].value}'

        assert 'История перемещений по стадиям' == sheet_names[4], \
            f"Expected История перемещений по стадиям, but got {sheet_names[4]}"
        assert sheet_fifth['A2'].value == 'Номер', f'Expected Номер, but got {sheet_fifth["A2"].value}'
        assert sheet_fifth['A3'].value == model_task.number, \
            f'Expected {model_task.number}, but got {sheet_fifth["A3"].value}'
        assert sheet_fifth['B2'].value == 'Стадия заявки', f'Expected Стадия заявки, but got {sheet_fifth["B2"].value}'
        assert sheet_fifth['B3'].value == model_task.taskStage.name, \
            f'Expected {model_task.taskStage.name}, but got {sheet_fifth["B3"].value}'
        assert sheet_fifth['C2'].value == 'Дата перехода', \
            f'Expected Дата перехода, but got {sheet_fifth["C2"].value}'
        # assert sheet_fifth['C3'].value == '', f'Expected , but got {sheet_fifth["C3"].value}'
        assert sheet_fifth['D2'].value == 'ФИО', f'Expected ФИО, but got {sheet_fifth["D2"].value}'
        assert sheet_fifth['D3'].value == model_task.requestedBy.firstName, \
            f'Expected {model_task.requestedBy.firstName}, but got {sheet_fifth["D3"].value}'

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Заявки.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully export extended list task by task number.')

    @allure.step("Exports the extended task list taking into account the specified filters by task number, V2.")
    def get_extended_export_task_by_task_number_v2(
            self, model_task, model_contact, model_asset, model_district
    ):
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_list_extended_v2_tasks_all_filters_by_task_id_endpoint(model_task.number),
            headers=self.headers.export_header(get_token())
        )
        end = time.time()
        logger.info(response.headers)
        self.attach_response_headers(response.headers)
        data_response = self.response_content(response)
        self.attach_time(start, end)
        self.attach_url(response.request.url)
        assert response.status_code == HTTPStatus.OK, \
            f'Expected status code {HTTPStatus.OK}, but got {response.status_code}, {data_response}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        sheet = workbook.worksheets[0]
        sheet_second = workbook.worksheets[1]
        sheet_third = workbook.worksheets[2]
        sheet_fourth = workbook.worksheets[3]
        sheet_fifth = workbook.worksheets[4]
        sheet_names = workbook.sheetnames

        assert 'Заявки' == sheet_names[0], f"Expected Заявки, but got {sheet_names[0]}"
        assert sheet['A2'].value == 'Номер', f'Expected Номер, but got {sheet['A2'].value}'
        assert sheet['A3'].value == model_task.number, f'Expected {model_task.number}, but got {sheet['A3'].value}'
        assert sheet['B2'].value == 'Родительская заявка', f'Expected Родительская заявка, but got {sheet['B2'].value}'
        assert sheet['B3'].value == getattr(getattr(model_task, 'parent', None), 'name', ''), \
            f'Expected {getattr(getattr(model_task, 'parent', None), 'name', '')}, but got {sheet['B3'].value}'
        assert sheet['C2'].value == 'Тип заявки', f'Expected Тип заявки, but got {sheet['C2'].value}'
        assert sheet['C3'].value == model_task.taskType.name.strip(), \
            f'Expected {model_task.taskType.name.strip()}, but got {sheet['C3'].value}.'
        assert sheet['D2'].value == 'Описание заявки', f'Expected Описание заявки, but got {sheet['D2'].value}'
        assert sheet['D3'].value == model_task.notes.strip(), \
            f'Expected {model_task.notes.strip()}, but got {sheet['D3'].value}'

        assert sheet['E2'].value == 'Заказчик', f'Expected Заказчик, but got {sheet['E2'].value}'
        assert sheet['E3'].value == model_task.company.name.strip(), \
            f'Expected {model_task.company.name.strip()}, but got {sheet['E3'].value}'
        assert sheet['F2'].value == 'Оборудование', f'Expected Оборудование, but got {sheet['F2'].value}'
        assert sheet['F3'].value == model_task.asset.name.strip(), \
            f'Expected {model_task.asset.name.strip()}, but got {sheet['F3'].value}'
        assert sheet['G2'].value == 'Серийный номер объекта', \
            f'Expected Серийный номер объекта, but got {sheet['G2'].value}'
        assert sheet['H2'].value == 'Адрес', f'Expected Адрес, but got {sheet['H2'].value}'
        assert sheet['H3'].value == model_task.location.address, \
            f'Expected {model_task.location.address}, but got {sheet['H3'].value}'
        assert sheet['I2'].value == 'Описание объекта', f'Expected Описание объекта, but got {sheet['I2'].value}'
        assert sheet['I3'].value == model_task.location.description.strip(), \
            f'Expected {model_task.location.description.strip()}, but got {sheet['I3'].value}'

        latitude, longitude = model_task.location.coordinate.split(":")

        latitude, longitude = float(latitude), float(longitude)

        assert sheet['J2'].value == 'Широта', f'Expected Широта, but got {sheet['J2'].value}'
        assert sheet['J3'].value == latitude, f'Expected {latitude}, but got {sheet['J3'].value}'
        assert sheet['K2'].value == 'Долгота', f'Expected Долгота, but got {sheet['K2'].value}'
        assert sheet['K3'].value == longitude, f'Expected {longitude}, but got {sheet['K3'].value}'
        assert sheet['L2'].value == 'Часовой пояс', f'Expected Часовой пояс, but got {sheet['L2'].value}'
        assert sheet['L3'].value == model_task.location.timeZone.name.strip(), \
            f'Expected {model_task.location.timeZone.name.strip()}, but got {sheet['L3'].value}'
        assert sheet['M2'].value == 'Страна', f'Expected Страна, but got {sheet["M2"].value}'
        assert sheet['M3'].value == model_task.location.country.name, \
            f'Expected {model_task.location.country.name}, but got {sheet["M3"].value}'
        assert sheet['N2'].value == 'Вид работ', f'Expected Вид работ, but got {sheet["N2"].value}'
        assert sheet['N3'].value == model_task.workType.name.strip(), \
            f'Expected {model_task.workType.name.strip()}, but got {sheet["N3"].value}'

        assert sheet['O2'].value == 'Стадия', f'Expected Стадия, but got {sheet["O2"].value}'
        assert sheet['O3'].value == model_task.taskStage.name, \
            f'Expected {model_task.taskStage.name}, but got {sheet["O3"].value}'
        assert sheet['P2'].value == 'Статус', f'Expected Статус, but got {sheet["P2"].value}'
        assert sheet['P3'].value == model_task.taskStatus.name, \
            f'Expected {model_task.taskStatus.name}, but got {sheet["P3"].value}'

        assert sheet['Q2'].value == 'Актуальность', f'Expected Актуальность, but got {sheet["Q2"].value}'
        # assert sheet['Q3'].value == '', f'Expected , but got {sheet["Q3"].value}'
        assert sheet['R2'].value == 'Критичность', f'Expected Критичность, but got {sheet["R2"].value}'
        assert sheet['R3'].value == model_task.actualCriticality.name, \
            f'Expected {model_task.actualCriticality.name}, but got {sheet["R3"].value}'
        assert sheet['S2'].value == 'Крайний срок закрытия', \
            f'Expected Крайний срок закрытия, but got {sheet["S2"].value}'
        # assert sheet['S3'].value == '', f'Expected , but got {sheet["S3"].value}'
        assert sheet['T2'].value == 'Исполнитель', f'Expected Исполнитель, but got {sheet["T2"].value}'
        assigned_to = ",".join(
            f"{stuff.lastName} {stuff.firstName} {stuff.middleName if stuff.middleName else ''}".strip()
            for stuff in model_task.listAssignedTo
        )
        assert sheet['T3'].value == assigned_to, f'Expected {assigned_to}, but got {sheet["T3"].value}'

        assert sheet['U2'].value == 'Дата назначения', f'Expected Дата назначения, but got {sheet["U2"].value}'
        # assert sheet['U3'].value == '', f'Expected , but got {sheet["U3"].value}'
        assert sheet['V2'].value == 'Назначено С', f'Expected Назначено С, but got {sheet["V2"].value}'
        # assert sheet['V3'].value == '', f'Expected , but got {sheet["V3"].value}'
        assert sheet['W2'].value == 'Назначено По', f'Expected Назначено По, but got {sheet["W2"].value}'
        # assert sheet['W3'].value == '', f'Expected , but got {sheet["W3"].value}'

        assert sheet['X2'].value == 'Контактное лицо', f'Expected Контактное лицо, but got {sheet["X2"].value}'
        assert sheet['X3'].value == model_contact.fullName.strip(), \
            f'Expected {model_contact.fullName.strip()}, but got {sheet["X3"].value}'

        assert sheet['Y2'].value == 'Электронная почта', f'Expected Электронная почта, but got {sheet["Y2"].value}'
        assert sheet['Y3'].value == model_contact.email.strip(), \
            f'Expected {model_contact.email.strip()}, but got {sheet["Y3"].value}'

        assert sheet['Z2'].value == 'Телефон', f'Expected Телефон, but got {sheet["Z2"].value}'
        assert sheet['Z3'].value == model_contact.phone.strip(), \
            f'Expected {model_contact.phone.strip()}, but got {sheet["Z3"].value}'

        assert sheet['AA2'].value == 'Обращение от', f'Expected Обращение от, but got {sheet["AA2"].value}'
        assert sheet['AA3'].value == model_task.requestedBy.firstName.strip(), \
            f'Expected {model_task.requestedBy.firstName.strip()}, but got {sheet["AA3"].value}'

        assert sheet['AB2'].value == 'Ответственный за объект', \
            f'Expected Ответственный за объект, but got {sheet["AB2"].value}'
        responsible_person = (f"{model_asset.responsiblePerson.lastName} "
                              f"{model_asset.responsiblePerson.firstName} "
                              f"{model_asset.responsiblePerson.middleName if model_asset.responsiblePerson.middleName else ''}".strip()
                              )
        assert sheet['AB3'].value == responsible_person, \
            f'Expected {responsible_person}, but got {sheet["AB3"].value}'

        assert sheet['AC2'].value == 'Название участка', f'Expected Название участка, but got {sheet["AC2"].value}'

        district = model_district.root[next(iter(model_district.root))].name.strip()
        assert sheet['AC3'].value == district, f'Expected {district}, but got {sheet["AC3"].value}'

        assert sheet['AD2'].value == 'Класс объекта', f'Expected Класс объекта, but got {sheet["AD2"].value}'
        assert sheet['AD3'].value == model_asset.assetClass.name.strip(), \
            f'Expected {model_asset.assetClass.name.strip()}, but got {sheet["AD3"].value}'

        assert sheet['AE2'].value == 'Тип объекта', f'Expected Тип объекта, but got {sheet["AE2"].value}'
        assert sheet['AE3'].value == model_asset.assetType.name.strip(), \
            f'Expected {model_asset.assetType.name.strip()}, but got {sheet["AE3"].value}'

        assert sheet['AF2'].value == 'Неисправность обнаружена', \
            f'Expected Неисправность обнаружена, but got {sheet["AF2"].value}'
        # assert sheet['AF3'].value == '', f'Expected , but got {sheet["AF3"].value}'

        assert sheet['AG2'].value == 'Выполнена', f'Expected Выполнена, but got {sheet["AG2"].value}'
        # assert sheet['AG3'].value == '', f'Expected , but got {sheet["AG3"].value}'

        assert sheet['AH2'].value == 'Дата создания', f'Expected Дата создания, but got {sheet["AH2"].value}'
        # assert sheet['AH3'].value == '', f'Expected , but got {sheet["AH3"].value}'

        assert sheet['AI2'].value == 'Закрыта', f'Expected Закрыта, but got {sheet["AI2"].value}'
        # assert sheet['AI3'].value == '', f'Expected , but got {sheet["AI3"].value}'

        assert sheet['AJ2'].value == 'Объект посещен', f'Expected Объект посещен, but got {sheet["AJ2"].value}'
        # assert sheet['AJ3'].value == '', f'Expected , but got {sheet["AJ3"].value}'

        assert sheet['AK2'].value == 'Метод подачи заявки', \
            f'Expected Метод подачи заявки, but got {sheet["AK2"].value}'
        assert sheet['AK3'].value == model_task.requestMethod.name.strip(), \
            f'Expected {model_task.requestMethod.name.strip()}, but got {sheet["AK3"].value}'

        assert sheet['AL2'].value == 'Оценочные трудозатраты', \
            f'Expected Оценочные трудозатраты, but got {sheet["AL2"].value}'
        # assert sheet['AL3'].value == '', f'Expected , but got {sheet["AL3"].value}'

        assert sheet['AM2'].value == 'Оценочная стоимость', \
            f'Expected Оценочная стоимость, but got {sheet["AM2"].value}'
        # assert sheet['AM3'].value == '', f'Expected , but got {sheet["AM3"].value}'

        assert sheet['AN2'].value == 'Фактические трудозатраты', \
            f'Expected Фактические трудозатраты, but got {sheet["AN2"].value}'
        # assert sheet['AN3'].value == '', f'Expected , but got {sheet["AN3"].value}'

        assert sheet['AO2'].value == 'Фактическая стоимость', \
            f'Expected Фактическая стоимость, but got {sheet["AO2"].value}'
        # assert sheet['AO3'].value == '', f'Expected , but got {sheet["AO3"].value}'

        assert sheet['AP2'].value.strip() == 'Шаблоны', f'Expected Шаблоны, but got {sheet["AP2"].value}'
        # assert sheet['AP3'].value == '', f'Expected , but got {sheet["AP3"].value}'

        assert 'Выполненные работы' == sheet_names[1], f"Expected Выполненные работы, but got {sheet_names[1]}"
        assert sheet_second['A2'].value == 'Номер', f'Expected Номер, but got {sheet_second["A2"].value}'
        # assert sheet_second['A3'].value == '', f'Expected , but got {sheet_second["A3"].value}'
        assert sheet_second['B2'].value == 'Вид работ', f'Expected Вид работ, but got {sheet_second["B2"].value}'
        # assert sheet_second['B3'].value == '', f'Expected , but got {sheet_second["B3"].value}'
        assert sheet_second['C2'].value == 'Стоимость вида работ', \
            f'Expected Стоимость вида работ, but got {sheet_second["C2"].value}'
        # assert sheet_second['C3'].value == '', f'Expected , but got {sheet_second["C3"].value}'
        assert sheet_second['D2'].value == 'Оборудование', f'Expected Оборудование, but got {sheet_second["D2"].value}'
        # assert sheet_second['D3'].value == '', f'Expected , but got {sheet_second["D3"].value}'
        assert sheet_second['E2'].value == 'Выполнил', f'Expected Выполнил, but got {sheet_second["E2"].value}'
        # assert sheet_second['E3'].value == '', f'Expected , but got {sheet_second["E3"].value}'
        assert sheet_second['F2'].value == 'Ставка сотрудника', \
            f'Expected Ставка сотрудника, but got {sheet_second["F2"].value}'
        # assert sheet_second['F3'].value == '', f'Expected , but got {sheet_second["F3"].value}'
        assert sheet_second['G2'].value == 'Начало работ', f'Expected Начало работ, but got {sheet_second["G2"].value}'
        # assert sheet_second['G3'].value == '', f'Expected , but got {sheet_second["G3"].value}'
        assert sheet_second['H2'].value == 'Окончание работ', \
            f'Expected Окончание работ, but got {sheet_second["H2"].value}'
        # assert sheet_second['H3'].value == '', f'Expected , but got {sheet_second["H3"].value}'
        assert sheet_second['I2'].value == 'Объем работ', f'Expected Объем работ, but got {sheet_second["I2"].value}'
        # assert sheet_second['I3'].value == '', f'Expected , but got {sheet_second["I3"].value}'
        assert sheet_second['J2'].value == 'Единица измерения', \
            f'Expected Единица измерения, but got {sheet_second["J2"].value}'
        # assert sheet_second['J3'].value == '', f'Expected , but got {sheet_second["J3"].value}'
        assert sheet_second['K2'].value == 'Описание', f'Expected Описание, but got {sheet_second["K2"].value}'
        # assert sheet_second['K3'].value == '', f'Expected , but got {sheet_second["K3"].value}'

        assert 'Сообщения' == sheet_names[2], f"Expected Сообщения, but got {sheet_names[2]}"
        assert sheet_third['A2'].value == 'Номер', f'Expected Номер, but got {sheet_third["A2"].value}'
        # assert sheet_third['A3'].value == '', f'Expected , but got {sheet_third["A3"].value}'
        assert sheet_third['B2'].value == 'Оправитель', f'Expected Оправитель, but got {sheet_third["B2"].value}'
        # assert sheet_third['B3'].value == '', f'Expected , but got {sheet_third["B3"].value}'
        assert sheet_third['C2'].value == 'Сообщение', \
            f'Expected Сообщение, but got {sheet_third["C2"].value}'
        # assert sheet_third['C3'].value == '', f'Expected , but got {sheet_third["C3"].value}'
        assert sheet_third['D2'].value == 'Отправлено', f'Expected Отправлено, but got {sheet_third["D2"].value}'
        # assert sheet_third['D3'].value == '', f'Expected , but got {sheet_third["D3"].value}'

        assert 'Чек листы' == sheet_names[3], f"Expected Чек листы, but got {sheet_names[3]}"
        assert sheet_fourth['A2'].value == 'Номер', f'Expected Номер, but got {sheet_fourth["A2"].value}'
        # assert sheet_fourth['A3'].value == '', f'Expected , but got {sheet_fourth["A3"].value}'
        assert sheet_fourth['B2'].value == 'Чек лист', f'Expected Чек лист, but got {sheet_fourth["B2"].value}'
        # assert sheet_fourth['B3'].value == '', f'Expected , but got {sheet_fourth["B3"].value}'
        assert sheet_fourth['C2'].value == 'Пункт чек листа', \
            f'Expected Пункт чек листа, but got {sheet_fourth["C2"].value}'
        # assert sheet_fourth['C3'].value == '', f'Expected , but got {sheet_fourth["C3"].value}'
        assert sheet_fourth['D2'].value == 'Отмечено', f'Expected Отмечено, but got {sheet_fourth["D2"].value}'
        # assert sheet_fourth['D3'].value == '', f'Expected , but got {sheet_fourth["D3"].value}'
        assert sheet_fourth['E2'].value == 'Значение', f'Expected Значение, but got {sheet_fourth["E2"].value}'
        # assert sheet_fourth['E3'].value == '', f'Expected , but got {sheet_fourth["E3"].value}'
        assert sheet_fourth['F2'].value == 'Обработано', \
            f'Expected Обработано, but got {sheet_fourth["F2"].value}'
        # assert sheet_fourth['F3'].value == '', f'Expected , but got {sheet_fourth["F3"].value}'

        assert 'История перемещений по стадиям' == sheet_names[4], \
            f"Expected История перемещений по стадиям, but got {sheet_names[4]}"
        assert sheet_fifth['A2'].value == 'Номер', f'Expected Номер, but got {sheet_fifth["A2"].value}'
        assert sheet_fifth['A3'].value == model_task.number, \
            f'Expected {model_task.number}, but got {sheet_fifth["A3"].value}'
        assert sheet_fifth['B2'].value == 'Стадия заявки', f'Expected Стадия заявки, but got {sheet_fifth["B2"].value}'
        assert sheet_fifth['B3'].value == model_task.taskStage.name, \
            f'Expected {model_task.taskStage.name}, but got {sheet_fifth["B3"].value}'
        assert sheet_fifth['C2'].value == 'Дата перехода', \
            f'Expected Дата перехода, but got {sheet_fifth["C2"].value}'
        # assert sheet_fifth['C3'].value == '', f'Expected , but got {sheet_fifth["C3"].value}'
        assert sheet_fifth['D2'].value == 'ФИО', f'Expected ФИО, but got {sheet_fifth["D2"].value}'
        assert sheet_fifth['D3'].value == model_task.requestedBy.firstName, \
            f'Expected {model_task.requestedBy.firstName}, but got {sheet_fifth["D3"].value}'

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Заявки.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully export extended list task by task number, V2.')
