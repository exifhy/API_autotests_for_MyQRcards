from urllib import parse
import allure
import requests
from loguru import logger
from requests import JSONDecodeError
from utils.helper import Helper
from services.export.material_consumption.payloads import Payloads
from services.export.material_consumption.endpoints import Endpoints
from config.headers import Headers
from services.export.material_consumption.models.export_material_consumption_model import *
import time
from http import HTTPStatus
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from io import BytesIO


load_dotenv()
API_TOKEN = os.getenv('API_TOKEN')


class ExportMaterialConsumptionAPI(Helper):

    def __init__(self):
        super().__init__()
        self.payloads = Payloads()
        self.endpoints = Endpoints()
        self.headers = Headers()

    @allure.step("Exports a list of material costs.")
    def get_export_list_material_costs(self):
        params = {
            "noData": True
        }
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_material_consumption_endpoint, params=params,
            headers=self.headers.export_header(API_TOKEN)
        )
        end = time.time()
        logger.info(response.headers)
        self.attach_time(start, end)
        self.attach_url(response.request.url)
        try:
            logger.warning(response.json())
        except JSONDecodeError:
            logger.warning("Received response is not a valid JSON")
        assert response.status_code == HTTPStatus.OK, f'{response.status_code}, {response.json()}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        # СОХРАНЕНИЕ ЛОКАЛЬНО EXCEL
        # output_file_path = "exported_data.xlsx"
        # workbook.save(output_file_path)

        sheet = workbook.active
        sheet_name = workbook.sheetnames

        assert 'Расходы материалов' in sheet_name
        assert sheet['B3'].value == 'Номенклатура материала', f'Expected Номенклатура материала, but got {sheet['B3'].value}'
        assert sheet['C3'].value == 'Код номенклатуры материала', f'Expected Код номенклатуры материала, but got {sheet['C3'].value}'
        assert sheet['D3'].value == 'Стоимость (руб)', f'Expected Стоимость (руб), but got {sheet['D3'].value}'
        assert sheet['E3'].value == 'Расходы', f'Expected Расходы, but got {sheet['E3'].value}'
        assert sheet['F3'].value == 'Итоговая стоимость', f'Expected Итоговая стоимость, but got {sheet['F3'].value}'
        assert sheet['H3'].value == 'Название склада', f'Expected Название склада, but got {sheet['H3'].value}'
        assert sheet['I3'].value == 'Код склада', f'Expected Код склада, but got {sheet['I3'].value}'
        assert sheet['K3'].value == 'Единица измерения', f'Expected Единица измерения, but got {sheet['K3'].value}'
        assert sheet['M3'].value == 'Номер заявки', f'Expected Номер заявки, but got {sheet['M3'].value}'
        assert sheet['O3'].value == 'Оборудование', f'Expected Оборудование, but got {sheet['O3'].value}'
        assert sheet['P3'].value == 'Дата расхода', f'Expected Дата расхода, but got {sheet['P3'].value}'
        assert sheet['Q3'].value == 'ФИО сотрудника, чей расход', f'Expected ФИО сотрудника, чей расход, but got {sheet['Q3'].value}'

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Расходы+материалов.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully export of list material consumption.')
