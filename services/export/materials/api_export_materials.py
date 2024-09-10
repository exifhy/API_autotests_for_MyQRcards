from urllib import parse
import allure
import requests
from loguru import logger
from requests import JSONDecodeError
from utils.helper import Helper
from services.export.materials.payloads import Payloads
from services.export.materials.endpoints import Endpoints
from config.headers import Headers
from services.export.materials.models.export_materials_model import *
import time
from http import HTTPStatus
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from io import BytesIO


load_dotenv()
API_TOKEN = os.getenv('API_TOKEN')


class ExportMaterialsAPI(Helper):

    def __init__(self):
        super().__init__()
        self.payloads = Payloads()
        self.endpoints = Endpoints()
        self.headers = Headers()

    @allure.step("Exports the list of materials.")
    def get_export_list_materials(self):
        params = {
            "noData": True
        }
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_materials_endpoint, params=params,
            headers=self.headers.export_header(API_TOKEN)
        )
        end = time.time()
        logger.info(response.headers)
        logger.warning(response.request.url)
        self.attach_time(start, end)
        try:
            logger.warning(response.json())
        except JSONDecodeError:
            logger.warning("Received response is not a valid JSON")
        assert response.status_code == HTTPStatus.OK, f'Status code {response.status_code}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        # СОХРАНЕНИЕ ЛОКАЛЬНО EXCEL
        # output_file_path = "exported_data.xlsx"
        # workbook.save(output_file_path)

        sheet = workbook.active
        sheet_name = workbook.sheetnames

        assert 'Материалы' in sheet_name
        assert sheet['B3'].value == 'Номенклатура материала', f'Expected Номенклатура материала, but got {sheet['B3'].value}'
        assert sheet['C3'].value == 'Код номенклатуры материала', f'Expected Код номенклатуры материала, but got {sheet['C3'].value}'
        assert sheet['D3'].value == 'Стоимость (руб)', f'Expected Стоимость (руб), but got {sheet['D3'].value}'
        assert sheet['E3'].value == 'Название склада', f'Expected Название склада, but got {sheet['E3'].value}'
        assert sheet['F3'].value == 'Код склада', f'Expected Код склада, but got {sheet['F3'].value}'
        assert sheet['G3'].value == 'Кол-во на складе', f'Expected Кол-во на складе, but got {sheet['G3'].value}'
        assert sheet['H3'].value == 'Единица измерения', f'Expected Единица измерения, but got {sheet['H3'].value}'

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Материалы.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully export of list materials.')
