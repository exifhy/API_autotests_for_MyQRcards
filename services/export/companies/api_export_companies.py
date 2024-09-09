from urllib import parse
import allure
import requests
from loguru import logger
from requests import JSONDecodeError
from utils.helper import Helper
from services.export.companies.payloads import Payloads
from services.export.companies.endpoints import Endpoints
from config.headers import Headers
from services.export.companies.models.export_companies_model import *
import time
from http import HTTPStatus
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from io import BytesIO


load_dotenv()
API_TOKEN = os.getenv('API_TOKEN')


class ExportCompaniesAPI(Helper):

    def __init__(self):
        super().__init__()
        self.payloads = Payloads()
        self.endpoints = Endpoints()
        self.headers = Headers()

    @allure.step("Exports the list of companies taking into account the specified filters.")
    def get_export_list_companies(self):
        # params = {
        #     "noData": False,
        #     "erpID": ""
        # }
        start = time.time()
        response = requests.get(
            url=self.endpoints.export_list_companies_endpoint,
            headers=self.headers.export_header(API_TOKEN)
        )
        end = time.time()
        logger.info(response.headers)
        logger.warning(response.request.url)
        self.attach_time(start, end)
        try:
            logger.warning(response.json())
        except JSONDecodeError:
            logger.warning("Received response is not a valid JSON")
        assert response.status_code == HTTPStatus.OK, f'Status code {response.status_code}'
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_stream = BytesIO(response.content)
        workbook = load_workbook(file_stream)

        # СОХРАНЕНИЕ ЛОКАЛЬНО EXCEL
        # output_file_path = "exported_data.xlsx"
        # workbook.save(output_file_path)

        sheet = workbook.active
        sheet_name = workbook.sheetnames

        assert 'Компании' in sheet_name
        assert sheet['B3'].value == 'Название*', f'Expected Название*, but got {sheet['B3'].value}'
        assert sheet['C3'].value == 'Полное наименование', f'Expected Полное наименование, but got {sheet['C3'].value}'
        assert sheet['D3'].value == 'Код', f'Expected Код, but got {sheet['D3'].value}'
        assert sheet['E3'].value == 'ERP ID', f'Expected ERP ID, but got {sheet['E3'].value}'
        assert sheet['F3'].value == 'Адрес сайта', f'Expected Адрес сайта, but got {sheet['F3'].value}'
        assert sheet['G3'].value == 'Телефон', f'Expected Телефон, but got {sheet['G3'].value}'
        assert sheet['H3'].value == 'Электронная почта', f'Expected Электронная почта, but got {sheet['H3'].value}'
        assert sheet['J3'].value == 'Тип контрагента*', f'Expected Тип контрагента*, but got {sheet['J3'].value}'
        assert sheet['K3'].value == 'ИНН', f'Expected ИНН, but got {sheet['K3'].value}'
        assert sheet['M3'].value == 'Заказчик', f'Expected Заказчик, but got {sheet['M3'].value}'
        assert sheet['O3'].value == 'Подрядчик', f'Expected Подрядчик, but got {sheet['O3'].value}'
        assert sheet['Q3'].value == 'Наша компания', f'Expected Наша компания, but got {sheet['Q3'].value}'

        logger.warning(parse.unquote(response.headers['Content-Disposition']))
        expected_filename = "Компании.xlsx"
        expected_content = f'attachment; filename="{expected_filename}"; filename*=UTF-8\'\'"{expected_filename}"'
        decoded_content = parse.unquote(response.headers['Content-Disposition'])
        assert expected_content in decoded_content, f'Expected {expected_content}, but got {decoded_content}'
        logger.info(f'Successfully export of list companies.')
